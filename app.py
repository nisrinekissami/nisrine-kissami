"""
app.py — V4 (English UI)
------------------------
Automated downtime diagnostic dashboard — Versigent.

Changes in V4 (vs V3):
  - Entire application translated to English (labels, charts, exports, PPT).
  - Navigation moved to the very TOP of the sidebar (first thing visible,
    before Data/Column mapping/Filters) so switching pages never requires
    scrolling down.
  - The selected Phase (Assembly / Cutting) is now shown as a clear banner
    on every page, so it's always obvious which phase the results are based
    on.
  - Fixed a real bug in the Excel charts: data labels were showing
    series name + category name + value all stacked together (unreadable).
    All native Excel charts now explicitly show ONLY the value.
  - Large equipment/task groups (e.g. "Others" bucket with 200+ machines) are
    now capped to their top N most impactful items on every chart (app AND
    Excel), with the full detail still available in the data table below —
    otherwise both Plotly and native Excel charts become unreadable.
  - Versigent logo embedded directly in every Excel export sheet (not just
    colors).

Run with:
    streamlit run app.py
"""

import math
import os
from io import BytesIO

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.utils import get_column_letter

from utils import (
    ACCENT_COLOR,
    VERSIGENT_ACTION_PLAN_COLUMNS,
    VERSIGENT_BLACK_HEX,
    VERSIGENT_ORANGE_HEX,
    add_bar_chart_excel,
    add_equipement_column,
    add_image_to_sheet,
    add_logo_to_sheet,
    add_phase_column,
    add_repartition_chart_excel_labeled,
    auto_detect_family_column,
    auto_detect_position_column,
    build_ishikawa_fig,
    build_pptx_report,
    build_versigent_action_plan,
    clean_data,
    compute_alerts,
    compute_detailed_breakdown,
    compute_group_totals,
    compute_pareto,
    compute_pareto_level2,
    compute_paynter,
    exclude_week,
    format_excel_sheet,
    generate_text_summary,
    load_data,
    predict_next_week,
    standardize_columns,
    summary_kpis,
)

EXCLUDED_WEEK = 24
MAX_HOME_CHART_ITEMS = 15     # equipment mini Pareto charts (Home page, per-equipment tabs)
MAX_DETAIL_CHART_ITEMS = 30   # Detailed Graphs task-level charts (table stays full)

# ---------------------------------------------------------------------------
# GENERAL CONFIG
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Automated Downtime Dashboard", layout="wide", page_icon="🛠️")

try:
    _base_dir = os.path.dirname(os.path.abspath(__file__))
except NameError:
    _base_dir = os.getcwd()
LOGO_PATH = os.path.join(_base_dir, "assets", "versigent_logo.png")
LOGO_BYTES = None
if os.path.exists(LOGO_PATH):
    with open(LOGO_PATH, "rb") as _f:
        LOGO_BYTES = _f.read()

VERSIGENT_PIE_COLORS = ["#CD7925", "#1A1A1A", "#E8A867", "#595959",
                         "#B5651D", "#BFBFBF", "#8C5A2B", "#404040"]
EQUIPMENT_ORDER = ["CM", "ROB", "Ultrasonic", "Others"]

st.markdown(
    """<style>
    /* Long KPI values (e.g. machine or cause names) were being clipped/invisible
       inside st.metric. Force wrapping and a slightly smaller font so the full
       text is always visible instead of being cut off. */
    [data-testid="stMetricValue"] {
        white-space: normal;
        overflow-wrap: break-word;
        line-height: 1.2;
        font-size: 1.35rem;
    }
    [data-testid="stMetric"] {
        overflow: visible;
    }
    </style>""",
    unsafe_allow_html=True,
)

DARK_BG = "#12141C"
DARK_BAR = "#4C78E8"
DARK_LINE = "#E03C3C"
DARK_ACCENT = "#FFC94A"
DARK_GRID = "#2A2D3A"


def bar_fig(data, x_col, y_col, title, color=VERSIGENT_ORANGE_HEX, height=360, tickangle=-45):
    """Standard dashboard bar chart: value always shown above each bar, never a
    'Total Result' bar."""
    fig = go.Figure()
    fig.add_bar(x=data[x_col], y=data[y_col], marker_color=color,
                text=data[y_col], textposition="outside")
    # Extra headroom above the tallest bar so its value label is never cut off
    # by the top of the plot area (this was making the highest bar's number
    # invisible on several Pareto charts).
    max_val = data[y_col].max() if not data.empty else 0
    fig.update_layout(title=title, height=height, xaxis_tickangle=tickangle,
                       showlegend=False, margin=dict(t=60, b=10, l=10, r=10),
                       yaxis=dict(range=[0, max_val * 1.18 if max_val else 1]))
    return fig


def dark_pareto_fig(data, x_col, value_col, cumul_col, title, height=520):
    """Classic Pareto (bars + cumulative % line) in dark theme, matching the
    'Pareto - Fault Code' reference screenshot. Values shown above bars."""
    fig = go.Figure()
    fig.add_bar(x=data[x_col], y=data[value_col], name="Duration (min)", marker_color=DARK_BAR,
                text=data[value_col], textposition="outside", textfont=dict(color="white"))
    fig.add_trace(go.Scatter(x=data[x_col], y=data[cumul_col], name="Cumulative %", yaxis="y2",
                              mode="lines+markers", line=dict(color=DARK_LINE, width=3),
                              marker=dict(color=DARK_LINE, size=7)))
    fig.add_hline(y=80, line_dash="dash", line_color="#888", yref="y2")
    # Extra headroom above the tallest bar so its value label (the highest bar
    # in the Pareto) is always visible instead of being clipped at the top.
    max_val = data[value_col].max() if not data.empty else 0
    fig.update_layout(
        title=dict(text=title, font=dict(color="white", size=20)),
        paper_bgcolor=DARK_BG, plot_bgcolor=DARK_BG,
        font=dict(color="white"),
        xaxis=dict(color="white", gridcolor=DARK_GRID, tickangle=-45),
        yaxis=dict(color="white", gridcolor=DARK_GRID, title="Total duration (min)",
                    range=[0, max_val * 1.18 if max_val else 1]),
        yaxis2=dict(title="Cumulative %", overlaying="y", side="right", range=[0, 100],
                     color="white", gridcolor=DARK_GRID),
        legend=dict(font=dict(color="white")),
        height=height, margin=dict(t=70, b=10, l=10, r=10),
    )
    return fig


def fig_to_png_bytes(fig):
    try:
        return fig.to_image(format="png", scale=2, width=1100, height=550)
    except Exception:
        return None


def df_download_button(label, df_, filename, key):
    buf = BytesIO()
    df_.to_excel(buf, index=False)
    st.download_button(label, buf.getvalue(), file_name=filename, key=key)


def equipement_data(df_, equip_name, group_col="Machine", agg="sum", top_n=None):
    sub = df_[df_["Equipement"] == equip_name]
    return compute_group_totals(sub, group_col=group_col, agg=agg, top_n=top_n)


def phase_banner(phase_label):
    st.markdown(
        f"""<div style="background-color:{VERSIGENT_ORANGE_HEX};color:white;padding:8px 16px;
        border-radius:6px;font-weight:600;display:inline-block;margin-bottom:10px;">
        Results based on: {phase_label} phase</div>""",
        unsafe_allow_html=True,
    )


# ---------------------------------------------------------------------------
# EXCEL EXPORTS — "ready to send" sheets (large legible text, Versigent
# colors and logo, nothing overlapping), used by the Home / Pareto of the
# week / Action Plan / Data / Detailed Graphs pages.
# ---------------------------------------------------------------------------

def _write_kpi_block(ws, kpis, resume_text, start_row=2):
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    labels = ["Number of failures", "Total downtime", "Average", "Most critical equipment", "Most frequent cause"]
    values = [kpis["nb_evenements"], f"{kpis['duree_totale_h']} h", f"{kpis['duree_moyenne_min']} min",
              kpis["machine_top"], kpis["cause_top"]]
    for i, (lab, val) in enumerate(zip(labels, values)):
        col = 2 + i * 2
        cell_lab = ws.cell(row=start_row, column=col, value=lab)
        cell_lab.font = Font(bold=True, color="FFFFFF", size=11)
        cell_lab.fill = PatternFill("solid", fgColor=VERSIGENT_ORANGE_HEX.replace("#", ""))
        cell_lab.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_lab.border = border
        ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=col + 1)
        cell_val = ws.cell(row=start_row + 1, column=col, value=val)
        cell_val.font = Font(bold=True, size=16, color="1A1A1A")
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        cell_val.border = border
        ws.merge_cells(start_row=start_row + 1, start_column=col, end_row=start_row + 1, end_column=col + 1)
    ws.row_dimensions[start_row + 1].height = 26

    resume_row = start_row + 3
    cell_title = ws.cell(row=resume_row, column=2, value="Automatic summary")
    cell_title.font = Font(bold=True, size=13, color=VERSIGENT_ORANGE_HEX.replace("#", ""))
    cell_resume = ws.cell(row=resume_row + 1, column=2, value=resume_text)
    cell_resume.font = Font(size=11, color="1A1A1A")
    cell_resume.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=resume_row + 1, start_column=2, end_row=resume_row + 4, end_column=9)
    return resume_row + 6


def _write_title_banner(ws, title, n_cols=14, logo_bytes=None):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=18, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=VERSIGENT_ORANGE_HEX.replace("#", ""))
    cell.alignment = Alignment(horizontal="center", vertical="center", indent=6 if logo_bytes else 0)
    ws.row_dimensions[1].height = 34
    if logo_bytes:
        add_logo_to_sheet(ws, logo_bytes, anchor="A1", height_px=32)


def _write_hidden_table(ws, data, start_row, start_col=20):
    """Writes a small data table (used to feed native charts) further out on
    the sheet, outside the main visible area."""
    for j, col_name in enumerate(data.columns):
        ws.cell(row=start_row, column=start_col + j, value=col_name)
    for i, row in enumerate(data.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            ws.cell(row=start_row + i, column=start_col + j, value=val)
    return start_col


def _write_home_sheet(writer, df, phase_label, week_label, kpis, resume_text,
                       data_cm, data_rob, data_us, data_others, sheet_name="Assembly Analysis"):
    home_title = "Global Analysis" if (phase_label.startswith("All") or phase_label == "Global") else f"{phase_label} Analysis"
    df.iloc[0:0].to_excel(writer, sheet_name=sheet_name, index=False, startrow=100)  # ensure sheet exists
    ws = writer.sheets[sheet_name]
    _write_title_banner(ws, home_title, logo_bytes=LOGO_BYTES)
    next_row = _write_kpi_block(ws, kpis, resume_text, start_row=3)

    for c in range(1, 15):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.sheet_view.showGridLines = False

    tables = [("CM", data_cm), ("ROB", data_rob), ("Ultrasonic", data_us), ("Others", data_others)]
    anchors = {}
    header_row = next_row
    col_cursor = 20
    for label, data in tables:
        if data.empty:
            continue
        _write_hidden_table(ws, data, header_row, start_col=col_cursor)
        anchors[label] = (header_row, col_cursor, len(data))
        col_cursor += 3

    chart_anchor_row = next_row
    if "CM" in anchors:
        r, c, n = anchors["CM"]
        add_bar_chart_excel(ws, header_row=r, n_rows=n, cat_col=c, val_col=c + 1,
                             title="Pareto — CM", anchor=f"B{chart_anchor_row}")
    if "ROB" in anchors:
        r, c, n = anchors["ROB"]
        add_bar_chart_excel(ws, header_row=r, n_rows=n, cat_col=c, val_col=c + 1,
                             title="Pareto — ROB", anchor=f"H{chart_anchor_row}")
    if "Ultrasonic" in anchors:
        r, c, n = anchors["Ultrasonic"]
        add_bar_chart_excel(ws, header_row=r, n_rows=n, cat_col=c, val_col=c + 1,
                             title="Pareto — Ultrasonic", anchor=f"B{chart_anchor_row + 18}")
    if "Others" in anchors:
        r, c, n = anchors["Others"]
        add_bar_chart_excel(ws, header_row=r, n_rows=n, cat_col=c, val_col=c + 1,
                             title="Pareto — Others", anchor=f"H{chart_anchor_row + 18}")

    pie_source = df[df["Equipement"].isin(["CM", "ROB", "Ultrasonic"])]
    if not pie_source.empty:
        pie_data = pie_source.groupby("Equipement")["Duree_min"].sum().reset_index()
        # Trié par valeur décroissante + couleurs par RANG (pas par catégorie),
        # exactement comme le camembert de l'application : la part la plus
        # importante est toujours rouge, la moyenne bleue, la plus petite verte
        # — quel que soit l'équipement (CM/ROB/Ultrasonic) concerné. Sans ce tri,
        # le classeur Excel utilisait la palette orange par défaut.
        pie_data = pie_data.sort_values("Duree_min", ascending=False).reset_index(drop=True)
        pie_row = header_row
        pie_col = col_cursor
        _write_hidden_table(ws, pie_data, pie_row, start_col=pie_col)
        add_repartition_chart_excel_labeled(ws, header_row=pie_row, n_rows=len(pie_data), cat_col=pie_col,
                                             val_col=pie_col + 1, title="Breakdown — ROB / Ultrasonic / CM",
                                             anchor=f"B{chart_anchor_row + 36}",
                                             colors=["E03C3C", "1E88E5", "2E7D32"])
        col_cursor += 3

    nb_visits = compute_group_totals(df, group_col="Machine", agg="count", top_n=MAX_HOME_CHART_ITEMS)
    if not nb_visits.empty:
        _write_hidden_table(ws, nb_visits, header_row, start_col=col_cursor)
        add_bar_chart_excel(ws, header_row=header_row, n_rows=len(nb_visits), cat_col=col_cursor,
                             val_col=col_cursor + 1, title="Pareto — Number of visits",
                             anchor=f"H{chart_anchor_row + 36}", color=VERSIGENT_BLACK_HEX.replace("#", ""),
                             value_axis_title="Number of failures")

    ws.sheet_properties.tabColor = VERSIGENT_ORANGE_HEX.replace("#", "")


def _write_pareto_week_sheet(writer, pareto_week_top10, week_label, sheet_name="2. Pareto - Top 10"):
    listing = pareto_week_top10.copy()
    listing.insert(0, "Rank", range(1, len(listing) + 1))
    listing = listing.rename(columns={"Machine": "Equipment", "Valeur": "Total duration (min)"})
    # Table shifted to the right (column J) so the Pareto chart sits on the LEFT,
    # same design/structure as the other sheets — a plain white sheet, no gridlines.
    table_startcol = 9  # column J (0-indexed)
    listing.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, startcol=table_startcol)
    ws = writer.sheets[sheet_name]
    _write_title_banner(ws, "Pareto - Top 10", n_cols=6, logo_bytes=LOGO_BYTES)
    format_excel_sheet(ws, n_cols=len(listing.columns), header_row=3, header_color=ACCENT_COLOR,
                        start_col=table_startcol + 1)
    for c in range(table_startcol + 1, table_startcol + 1 + len(listing.columns)):
        ws.column_dimensions[get_column_letter(c)].width = 24
    add_bar_chart_excel(ws, header_row=3, n_rows=len(listing), cat_col=table_startcol + 2,
                         val_col=table_startcol + 3, title="Pareto - Top 10",
                         anchor="B3", width=20, height=11)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = VERSIGENT_ORANGE_HEX.replace("#", "")


def _write_ishikawa_sheet(writer, pareto_week_top10, sheet_name="3. Ishikawa - Top 10"):
    """Feuille Excel dédiée au diagramme d'Ishikawa (arête de poisson) du Top 10 —
    diagramme inséré en image (PNG) à GAUCHE, table du Top 10 à droite, même
    style/structure (bannière orange, logo, pas de quadrillage) que les autres
    feuilles de l'export global."""
    listing = pareto_week_top10.copy()
    listing.insert(0, "Rank", range(1, len(listing) + 1))
    listing = listing.rename(columns={"Machine": "Equipment", "Valeur": "Total duration (min)"})
    table_startcol = 16  # table décalée à droite pour laisser la place à l'image à gauche
    listing.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, startcol=table_startcol)
    ws = writer.sheets[sheet_name]
    _write_title_banner(ws, "Ishikawa Diagram - Top 10", n_cols=14, logo_bytes=LOGO_BYTES)
    format_excel_sheet(ws, n_cols=len(listing.columns), header_row=3, header_color=ACCENT_COLOR,
                        start_col=table_startcol + 1)
    for c in range(table_startcol + 1, table_startcol + 1 + len(listing.columns)):
        ws.column_dimensions[get_column_letter(c)].width = 22

    total_top10 = pareto_week_top10["Valeur"].sum() if not pareto_week_top10.empty else 0
    fig_ishi = build_ishikawa_fig(
        pareto_week_top10, group_col="Machine", value_col="Valeur",
        effect_label=f"Total Downtime<br>Top 10<br>{total_top10:,.0f} min",
        title="Ishikawa Diagram - Top 10",
    )
    png_ishi = fig_to_png_bytes(fig_ishi)
    add_image_to_sheet(ws, png_ishi, anchor="B3", width_px=980, height_px=480)

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = VERSIGENT_ORANGE_HEX.replace("#", "")


def _write_action_plan_sheet(writer, plan, phase_label, sheet_name="Action Plan"):
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    plan.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
    ws = writer.sheets[sheet_name]
    _write_title_banner(ws, f"{phase_label} Action Plan", n_cols=len(VERSIGENT_ACTION_PLAN_COLUMNS), logo_bytes=LOGO_BYTES)

    header_row = 3
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, col_name in enumerate(VERSIGENT_ACTION_PLAN_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=j)
        cell.value = col_name
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=VERSIGENT_ORANGE_HEX.replace("#", ""))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(j)].width = 20
    ws.row_dimensions[header_row].height = 34

    for i in range(len(plan)):
        r = header_row + 1 + i
        ws.row_dimensions[r].height = 32
        for j in range(1, len(VERSIGENT_ACTION_PLAN_COLUMNS) + 1):
            cell = ws.cell(row=r, column=j)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
            cell.border = border
            cell.font = Font(size=10.5, color="1A1A1A")
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="FBF0E4")
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = VERSIGENT_ORANGE_HEX.replace("#", "")


def _write_data_sheet(writer, data, sheet_name="5. Data"):
    from openpyxl.styles import Alignment

    data = data.rename(columns={"Equipement": "Equipment"})
    data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
    ws = writer.sheets[sheet_name]
    _write_title_banner(ws, "Data — failure detail for the selected scope", n_cols=len(data.columns), logo_bytes=LOGO_BYTES)

    header_row = 3
    # max_width=55 (au lieu du plafond par défaut de 45) car cette feuille contient
    # des colonnes à texte libre potentiellement longues (Cause, Task). L'ancienne
    # version écrasait ensuite CETTE largeur automatique avec une largeur fixe de
    # 20 pour toutes les colonnes -> les textes longs étaient coupés (illisibles)
    # et les colonnes courtes (Duree_min, Phase...) étaient inutilement élargies.
    # Cette largeur automatique par colonne est maintenant conservée telle quelle.
    format_excel_sheet(ws, n_cols=len(data.columns), header_row=header_row, header_color=ACCENT_COLOR,
                        max_width=55)

    # Pour les colonnes à texte libre pouvant dépasser même ce plafond de 55
    # caractères (ex: descriptions de panne longues), on active le retour à la
    # ligne ET on agrandit automatiquement la hauteur de la ligne concernée —
    # ainsi rien n'est jamais coupé/invisible et l'opérateur n'a jamais besoin de
    # retoucher lui-même les colonnes/lignes avant d'envoyer le fichier.
    wrap_cols = [c for c in ["Machine", "Cause", "Task"] if c in data.columns]
    if wrap_cols:
        col_letters = {col: get_column_letter(list(data.columns).index(col) + 1) for col in wrap_cols}
        col_widths = {col: ws.column_dimensions[col_letters[col]].width or 20 for col in wrap_cols}
        for row_idx in range(header_row + 1, header_row + 1 + len(data)):
            max_lines = 1
            for col in wrap_cols:
                letter = col_letters[col]
                cell = ws[f"{letter}{row_idx}"]
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.value is not None:
                    width_chars = max(int(col_widths[col]), 5)
                    lines_needed = math.ceil(len(str(cell.value)) / width_chars)
                    max_lines = max(max_lines, lines_needed)
            if max_lines > 1:
                ws.row_dimensions[row_idx].height = 15 * max_lines

    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = VERSIGENT_ORANGE_HEX.replace("#", "")


def _write_detailed_graphs_workbook(writer, detail_data):
    for equip, full_data in detail_data.items():
        sheet_name = equip[:31]
        capped = full_data.head(MAX_DETAIL_CHART_ITEMS)
        table = full_data.rename(columns={"Machine": "Equipment", "Task": "Failure description", "Duree_min": "Duration (min)"})
        table.insert(0, "Label", table["Equipment"].astype(str) + " — " + table["Failure description"].astype(str))
        table.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
        ws = writer.sheets[sheet_name]
        note = (f"Chart shows the top {min(MAX_DETAIL_CHART_ITEMS, len(full_data))} of {len(full_data)} items "
                f"by duration — full detail in the table below.") if len(full_data) > MAX_DETAIL_CHART_ITEMS else \
               "Full detail shown below."
        _write_title_banner(ws, f"Detailed Graphs — {equip}", n_cols=4, logo_bytes=LOGO_BYTES)
        ws.cell(row=2, column=1, value=note)
        from openpyxl.styles import Font
        ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="595959")
        format_excel_sheet(ws, n_cols=len(table.columns), header_row=4, header_color=ACCENT_COLOR)
        for c, width in zip(range(1, 5), [45, 22, 32, 14]):
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.sheet_view.showGridLines = False
        chart_table = capped.rename(columns={"Machine": "Equipment", "Task": "Failure description", "Duree_min": "Duration (min)"})
        chart_label_col = 20
        _write_hidden_table(
            ws,
            pd.DataFrame({"Label": chart_table["Equipment"].astype(str) + " — " + chart_table["Failure description"].astype(str),
                          "Duration (min)": chart_table["Duration (min)"]}),
            start_row=4, start_col=chart_label_col,
        )
        add_bar_chart_excel(ws, header_row=4, n_rows=len(capped), cat_col=chart_label_col, val_col=chart_label_col + 1,
                             title=f"Detailed Graphs — {equip}", anchor="F4", width=24,
                             height=max(10, 0.4 * max(len(capped), 1)), value_axis_title="Duration (min)")


def _write_detailed_graphs_summary_sheet(writer, detail_data, sheet_name="6. Detailed Graphs"):
    """Consolidated, single-page version of Detailed Graphs (one sheet, not one
    per equipment) so the global export stays at 5 pages total. The full
    per-equipment detail remains available in the standalone Detailed Graphs
    export on its own page."""
    frames = []
    for equip, full_data in detail_data.items():
        t = full_data.copy()
        t.insert(0, "Equipment", equip)
        frames.append(t)
    combined = pd.concat(frames, ignore_index=True).sort_values("Duree_min", ascending=False) if frames else \
        pd.DataFrame(columns=["Equipment", "Machine", "Task", "Duree_min"])

    capped = combined.head(MAX_DETAIL_CHART_ITEMS)

    table = combined.rename(columns={"Task": "Failure description", "Duree_min": "Duration (min)"})
    table.insert(0, "Label", table["Equipment"].astype(str) + " — " + table["Machine"].astype(str))
    table.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
    ws = writer.sheets[sheet_name]
    note = (f"Chart shows the top {min(MAX_DETAIL_CHART_ITEMS, len(combined))} of {len(combined)} items by duration "
            f"across all equipment — full per-equipment detail is available in the standalone Detailed Graphs export.")
    _write_title_banner(ws, "Detailed Graphs — All Equipment", n_cols=5, logo_bytes=LOGO_BYTES)
    ws.cell(row=2, column=1, value=note)
    from openpyxl.styles import Font
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="595959")
    format_excel_sheet(ws, n_cols=len(table.columns), header_row=4, header_color=ACCENT_COLOR)
    for c, width in zip(range(1, len(table.columns) + 1), [45, 22, 22, 32, 14]):
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.sheet_view.showGridLines = False

    if not capped.empty:
        chart_label_col = 20
        chart_labels = pd.DataFrame({
            "Label": capped["Equipment"].astype(str) + " — " + capped["Machine"].astype(str),
            "Duration (min)": capped["Duree_min"],
        })
        _write_hidden_table(ws, chart_labels, start_row=4, start_col=chart_label_col)
        add_bar_chart_excel(ws, header_row=4, n_rows=len(capped), cat_col=chart_label_col, val_col=chart_label_col + 1,
                             title="Detailed Graphs — All Equipment", anchor="G4", width=24,
                             height=max(10, 0.4 * max(len(capped), 1)), value_axis_title="Duration (min)")
    ws.sheet_properties.tabColor = VERSIGENT_ORANGE_HEX.replace("#", "")


# ---------------------------------------------------------------------------
# HEADER
# ---------------------------------------------------------------------------
col_logo, col_title = st.columns([1, 5])
with col_logo:
    if LOGO_BYTES:
        st.image(LOGO_BYTES, use_container_width=True)
with col_title:
    st.title("🛠️ Automated Downtime Dashboard")
    st.caption(
        "Automatic replica of the Versigent Excel workbook (Assembly / Cutting Analysis / "
        "Fault Code Analysis / Pareto by Equipment / Detailed Graphs / Action Plan / Data) — "
        "no more rebuilding these charts by hand."
    )

with st.expander("ℹ️ How to use this dashboard (read this first)"):
    st.markdown("""
    1. Load one or more Excel/CSV failure files in the sidebar (one file per month for example —
       they're combined automatically), or check the **sample dataset** box.
    2. Check/fix the column mapping if needed (section 2).
    3. Choose the **Phase** (Assembly / Cutting) then the **Equipment** to analyze (section 3) —
       the list updates automatically based on the chosen phase. All charts automatically group
       equipment into 4 categories: **CM** (also includes Conveyors), **ROB**, **Ultrasonic**, and
       **Others** (Torque, DS20, UCAB...).
    4. Use the navigation buttons at the top of the sidebar to open each section (just like the
       Excel workbook). Clicking the same button twice returns to the home page.

    ℹ️ **Week 24** is automatically excluded from the analysis (data considered unreliable).
    """)

# ---------------------------------------------------------------------------
# NAVIGATION — kept at the very TOP of the sidebar, first thing visible,
# so switching pages never requires scrolling down past Data/Filters.
# ---------------------------------------------------------------------------
# The very first item's label follows the phase actually selected (Assembly /
# Cutting) — e.g. "Assembly Analysis" or "Cutting Analysis". The phase itself is
# only known further down the script (once the data is loaded), so we cache it
# in session_state and reuse it here; it defaults to "Assembly" before any file
# is loaded and updates on the next rerun once the real phase is known.
_cached_phase_for_nav = st.session_state.get("phase_label_cache", "Assembly")
if _cached_phase_for_nav.startswith("All") or _cached_phase_for_nav == "Global":
    _home_nav_label = "🏭 Global Analysis"
else:
    _home_nav_label = f"🏭 {_cached_phase_for_nav} Analysis"

NAV_ITEMS = [
    ("home", _home_nav_label),
    ("pareto_week", "📊 Pareto - Top 10"),
    ("ishikawa", "🐟 Ishikawa - Top 10"),
    ("action_plan", "🗂️ Action Plan"),
    ("detailed", "📈 Detailed Graphs"),
    ("fault_code", "🎯 Fault Code Analysis"),
    ("pareto_equip", "🔍 Pareto by Equipment"),
    ("by_date", "📅 View by Date"),
    ("trend", "📉 Trend & Alerts"),
    ("data", "🗄️ Data"),
    ("export", "📦 Export"),
]
if "page" not in st.session_state:
    st.session_state.page = "home"

st.sidebar.header("Navigation")
for key, label in NAV_ITEMS:
    active = st.session_state.page == key
    if st.sidebar.button(label, key=f"nav_{key}", use_container_width=True,
                          type="primary" if active else "secondary"):
        st.session_state.page = "home" if active else key
st.sidebar.divider()

# ---------------------------------------------------------------------------
# SIDEBAR: 1. DATA
# ---------------------------------------------------------------------------
st.sidebar.header("1. Data")

use_sample = st.sidebar.checkbox("Use the sample dataset (Versigent)", value=False)

uploaded_files = None
if not use_sample:
    uploaded_files = st.sidebar.file_uploader(
        "Load one or more failure files (.xlsx, .xls or .csv)",
        type=["xlsx", "xls", "csv"], accept_multiple_files=True,
        help="One file per month for example: they are automatically combined into a single "
             "analysis. Just reload the new files every month.",
    )

raw_df = None
if use_sample:
    try:
        raw_df = pd.read_excel("donnees_pannes_exemple.xlsx")
        st.sidebar.success("Sample dataset loaded.")
    except FileNotFoundError:
        st.sidebar.error(
            "The sample file \u00ab donnees_pannes_exemple.xlsx \u00bb was not found on the server. "
            "Uncheck the box above and load your own files."
        )
elif uploaded_files:
    try:
        frames = [load_data(f) for f in uploaded_files]
        raw_df = pd.concat(frames, ignore_index=True)
        st.sidebar.success(f"{len(uploaded_files)} file(s) loaded and combined: {len(raw_df)} rows in total.")
    except Exception as e:
        st.sidebar.error(f"Error reading the files: {e}")

if raw_df is None:
    st.info("👈 Load one or more data files, or check the sample box to get started.")
    st.stop()

# ---------------------------------------------------------------------------
# 2. COLUMN MAPPING
# ---------------------------------------------------------------------------
st.sidebar.header("2. Column Mapping")
cols = list(raw_df.columns)


def guess(candidates):
    for c in candidates:
        for col in cols:
            if c.lower() in str(col).lower():
                return col
    return None


def guessed_index(guess_result):
    return cols.index(guess_result) if guess_result in cols else 0


mapping_ui = {}
mapping_ui["Date"] = st.sidebar.selectbox("Date column", cols, index=guessed_index(guess(["date"])))
mapping_ui["Machine"] = st.sidebar.selectbox(
    "Machine column", cols,
    index=guessed_index(guess(["asset description", "machine", "equipment", "equipement", "asset"])),
    help="Column that identifies the exact piece of equipment that failed (e.g. 'Asset Description').",
)
mapping_ui["Cause"] = st.sidebar.selectbox(
    "Cause / Fault Code column", cols,
    index=guessed_index(guess(["fault code", "cause", "reason", "defect"])),
)
mapping_ui["Duree_min"] = st.sidebar.selectbox(
    "Duration column (minutes)", cols,
    index=guessed_index(guess(["total duration", "duration"])),
)

task_guess = guess(["task title", "task"])
use_task = st.sidebar.checkbox("My file has a Failure Title column (e.g. Task Title)", value=task_guess is not None)
if use_task:
    mapping_ui["Task"] = st.sidebar.selectbox("Failure title column", cols, index=guessed_index(task_guess))

completion_guess = guess(["completion details", "completion"])
use_completion = st.sidebar.checkbox(
    "My file has a Completion Details column",
    value=completion_guess is not None,
    help="Used as the problem description in the action plan.",
)
if use_completion:
    mapping_ui["Completion"] = st.sidebar.selectbox("Completion details column", cols, index=guessed_index(completion_guess))

engineer_guess = guess(["principal engineer", "engineer"])
use_engineer = st.sidebar.checkbox(
    "My file has a Responsible Engineer / Technician column (e.g. Principal Engineer)",
    value=engineer_guess is not None,
    help="Used to automatically fill in the Resp. field of the action plan.",
)
if use_engineer:
    mapping_ui["Engineer"] = st.sidebar.selectbox("Responsible engineer column", cols, index=guessed_index(engineer_guess))

family_guess = auto_detect_family_column(cols)
use_family = st.sidebar.checkbox(
    "My file has an Equipment / Machine Category column (e.g. Sub Description)",
    value=family_guess is not None,
    help="Used to automatically group equipment into 4 categories: CM (+ Conveyor), ROB, "
         "Ultrasonic, Others.",
)
if use_family:
    mapping_ui["Famille"] = st.sidebar.selectbox("Equipment / Category column", cols, index=guessed_index(family_guess))

position_guess = auto_detect_position_column(cols)
use_position = st.sidebar.checkbox(
    "My file has a hierarchical Position column (e.g. 'Building,Manufacturing Assembly_FA,...')",
    value=position_guess is not None,
    help="Used to automatically determine the Phase (Assembly / Cutting) of each failure.",
)
position_col_choice = None
if use_position:
    position_col_choice = st.sidebar.selectbox("Hierarchical Position column", cols, index=guessed_index(position_guess))

raw_position_series = raw_df[position_col_choice] if position_col_choice else None
inverse_mapping = {v: k for k, v in mapping_ui.items()}

try:
    df = standardize_columns(raw_df, inverse_mapping)
    df = clean_data(df)
    df = add_phase_column(df, raw_position_series)
    df = add_equipement_column(df, famille_col="Famille")
    df = exclude_week(df, week_number=EXCLUDED_WEEK, date_col="Date")
except Exception as e:
    st.error(f"Error preparing the data: {e}")
    st.stop()

if df.attrs.get("lignes_supprimees", 0) > 0:
    st.sidebar.warning(f"{df.attrs['lignes_supprimees']} invalid row(s) ignored (missing date or duration).")
if df.attrs.get("dates_corrigees", 0) > 0:
    st.sidebar.warning(
        f"🗓️ {df.attrs['dates_corrigees']} date(s) had day and month swapped — corrected automatically."
    )
st.sidebar.caption(f"ℹ️ Week {EXCLUDED_WEEK} is automatically excluded from the analysis.")

_colonnes_a_verifier = [("Cause", "Cause"), ("Machine", "Machine")]
if "Famille" in df.columns:
    _colonnes_a_verifier.append(("Famille", "Equipment / Category"))
for col_standard, label in _colonnes_a_verifier:
    if col_standard in df.columns and df[col_standard].nunique(dropna=True) <= 1:
        valeur_unique = df[col_standard].iloc[0] if not df.empty else "?"
        st.warning(
            f"⚠️ The column mapped to **{label}** has only one value across all failures "
            f"(\u00ab {valeur_unique} \u00bb). Check the column mapping in the sidebar."
        )

if df.empty:
    st.warning("No usable data after cleaning (and excluding week 24). Please check your file.")
    st.stop()

# ---------------------------------------------------------------------------
# 3. PHASE & EQUIPMENT (cascading selection)
# ---------------------------------------------------------------------------
st.sidebar.header("3. Phase & Equipment")

# Only Assembly / Cutting count as real phases for this selector — "Other" rows
# (position values that matched neither keyword) are never dropped, they just
# aren't offered as their own choice.
phases_presentes = [p for p in ["Assembly", "Cutting"] if p in df["Phase"].unique()] if "Phase" in df.columns else []

if len(phases_presentes) >= 2:
    # The file has BOTH phases -> let the user isolate one, and make it explicit
    # that "All" means the two combined (this is the case that used to be
    # ambiguous: "All" was shown with no indication of what it covered).
    phase_options = ["All (Assembly + Cutting)"] + phases_presentes
    _phase_choice = st.sidebar.selectbox(
        "Choose the phase", phase_options, index=0,
        help="Automatically detected from the Position column (Manufacturing "
             "Assembly_FA -> Assembly ; Cutting_CT / Die Center_DC / Lead Prep_LP -> "
             "Cutting). Pick Assembly or Cutting to restrict every graph and Pareto "
             "to that phase only, or keep 'All' to combine both.",
    )
    if _phase_choice.startswith("All"):
        selected_phase = "All"
        phase_label = "All (Assembly + Cutting)"
        phase_label_file = "All_Assembly_Cutting"
    else:
        selected_phase = _phase_choice
        phase_label = _phase_choice
        phase_label_file = _phase_choice
elif len(phases_presentes) == 1:
    # The file has only ONE phase -> fully automatic, nothing to choose.
    selected_phase = "All"  # every row already belongs to this one phase
    phase_label = phases_presentes[0]
    phase_label_file = phases_presentes[0]
    st.sidebar.success(
        f"✅ Phase auto-detected: **{phase_label}** — this file only contains "
        f"{phase_label} data, so every graph and Pareto below is {phase_label} only."
    )
else:
    # No Position column mapped / no phase could be inferred -> legacy behavior.
    selected_phase = "All"
    phase_label = "Global"
    phase_label_file = "Global"

df_phase = df if selected_phase == "All" else df[df["Phase"] == selected_phase]

# Cache the phase label so the navigation item built at the top of the sidebar
# (rendered before this point) can show "Assembly Analysis" / "Cutting Analysis"
# on the next rerun.
st.session_state["phase_label_cache"] = phase_label

equipements_dispo = [e for e in EQUIPMENT_ORDER if e in df_phase["Equipement"].unique()]
selected_equipements = st.sidebar.multiselect(
    "Choose the equipment", equipements_dispo, default=equipements_dispo,
    help="CM automatically includes Conveyors. Others groups the less frequent equipment "
         "(Torque, DS20, UCAB...).",
)
df_scope = df_phase[df_phase["Equipement"].isin(selected_equipements)] if selected_equipements else df_phase.iloc[0:0]

# ---------------------------------------------------------------------------
# 4. FILTERS (machine + period)
# ---------------------------------------------------------------------------
st.sidebar.header("4. Filters")
machines = sorted(df_scope["Machine"].unique().tolist())
selected_machines = st.sidebar.multiselect("Filter by machine", machines, default=machines)
df_scope = df_scope[df_scope["Machine"].isin(selected_machines)]

date_min, date_max = df_scope["Date"].min(), df_scope["Date"].max()
if pd.notna(date_min) and pd.notna(date_max):
    date_range = st.sidebar.date_input("Period", value=(date_min.date(), date_max.date()))
    if len(date_range) == 2:
        df_scope = df_scope[(df_scope["Date"] >= pd.to_datetime(date_range[0])) &
                             (df_scope["Date"] <= pd.to_datetime(date_range[1]))]

if df_scope.empty:
    st.warning("No data for this Phase / Equipment / Filters combination. Widen the selection in the sidebar.")
    st.stop()

df = df_scope  # from here on, `df` = selected scope (phase + equipment + filters)

week_label = ""
if df["Date"].notna().any():
    iso = df["Date"].dropna().dt.isocalendar()
    week_num = iso["week"].mode().iloc[0]
    week_label = f"week {week_num}"
completion_col = "Completion" if "Completion" in df.columns else None
engineer_col = "Engineer" if "Engineer" in df.columns else None

page = st.session_state.page

# ---------------------------------------------------------------------------
# PAGE: {PHASE} ANALYSIS (home) — e.g. "Assembly Analysis" / "Cutting Analysis"
# ---------------------------------------------------------------------------
if page == "home":
    if phase_label.startswith("All") or phase_label == "Global":
        home_page_title = "Global Analysis"
    else:
        home_page_title = f"{phase_label} Analysis"

    phase_banner(phase_label)
    st.header(f"🏭 {home_page_title}")
    kpis = summary_kpis(df)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Number of failures", kpis["nb_evenements"])
    c2.metric("Total downtime", f"{kpis['duree_totale_h']} h")
    c3.metric("Average", f"{kpis['duree_moyenne_min']} min", help="Total downtime divided by the number of failures.")
    c4.metric("Most critical equipment", kpis["machine_top"])
    c5.metric("Most frequent cause", kpis["cause_top"])

    pareto1_preview = compute_pareto(df, group_col="Cause")
    resume_text = generate_text_summary(df, pareto1_preview, group_col="Cause")
    st.info(f"📝 **Automatic summary** (ready to copy into your report):\n\n{resume_text}")

    st.divider()
    st.header(f"🏭 {phase_label} Analysis")
    st.caption("CM also includes Conveyors. « Others » groups all less frequent equipment.")

    data_cm = equipement_data(df, "CM", top_n=MAX_HOME_CHART_ITEMS)
    data_rob = equipement_data(df, "ROB", top_n=MAX_HOME_CHART_ITEMS)
    data_us = equipement_data(df, "Ultrasonic", top_n=MAX_HOME_CHART_ITEMS)
    data_others = equipement_data(df, "Others", top_n=MAX_HOME_CHART_ITEMS)
    full_others_count = len(equipement_data(df, "Others"))

    r1c1, r1c2 = st.columns(2)
    r1c1.plotly_chart(bar_fig(data_cm, "Machine", "Valeur", "Pareto — CM"), use_container_width=True)
    r1c2.plotly_chart(bar_fig(data_rob, "Machine", "Valeur", "Pareto — ROB"), use_container_width=True)

    r2c1, r2c2 = st.columns(2)
    r2c1.plotly_chart(bar_fig(data_us, "Machine", "Valeur", "Pareto — Ultrasonic"), use_container_width=True)
    with r2c2:
        title_others = "Pareto — Others"
        if full_others_count > MAX_HOME_CHART_ITEMS:
            st.caption(f"Showing top {MAX_HOME_CHART_ITEMS} of {full_others_count} pieces of equipment (see Data tab for full detail).")
        st.plotly_chart(bar_fig(data_others, "Machine", "Valeur", title_others), use_container_width=True)

    r3c1, r3c2 = st.columns(2)
    with r3c1:
        pie_source = df[df["Equipement"].isin(["CM", "ROB", "Ultrasonic"])]
        if not pie_source.empty:
            pie_data = pie_source.groupby("Equipement")["Duree_min"].sum().reindex(
                [e for e in ["CM", "ROB", "Ultrasonic"] if e in pie_source["Equipement"].unique()]).reset_index()
            # Color by value rank rather than by category: highest share = red,
            # middle = blue, lowest = green — regardless of which equipment
            # (CM/ROB/Ultrasonic) ends up in which rank.
            _rank_colors = ["#E03C3C", "#1E88E5", "#2E7D32"]
            _sorted_cats = pie_data.sort_values("Duree_min", ascending=False)["Equipement"].tolist()
            _color_map = {cat: _rank_colors[i] for i, cat in enumerate(_sorted_cats)}
            fig_pie = px.pie(pie_data, names="Equipement", values="Duree_min", hole=0,
                              title="Breakdown — ROB / Ultrasonic / CM",
                              color="Equipement", color_discrete_map=_color_map)
            # automargin=True laisse Plotly agrandir automatiquement les marges
            # nécessaires pour que les étiquettes extérieures (nom + %) ne
            # soient jamais coupées par le bord du graphique (ex: "Ultrasonic"
            # tronqué à gauche) ; marges de base généreuses en complément.
            fig_pie.update_traces(textposition="outside", textinfo="percent+label", automargin=True)
            fig_pie.update_layout(height=430, margin=dict(t=60, b=40, l=90, r=90))
            st.plotly_chart(fig_pie, use_container_width=True)
        else:
            st.info("No CM / ROB / Ultrasonic data in this scope.")
    with r3c2:
        nb_visits = compute_group_totals(df, group_col="Machine", agg="count", top_n=MAX_HOME_CHART_ITEMS)
        st.plotly_chart(bar_fig(nb_visits, "Machine", "Valeur", "Pareto — Number of visits", color="#8C8C8C"),
                         use_container_width=True)

    st.divider()
    st.subheader("📦 Export this page")
    buf_home = BytesIO()
    with pd.ExcelWriter(buf_home, engine="openpyxl") as writer:
        _write_home_sheet(writer, df, phase_label, week_label, kpis, resume_text,
                           data_cm, data_rob, data_us, data_others)
    st.download_button(
        f"📦 Download {home_page_title} (Excel, ready to send)",
        buf_home.getvalue(),
        file_name=f"{home_page_title.replace(' ', '_')}_{phase_label_file}_{pd.Timestamp.today().date()}.xlsx",
        key="dl_home_xlsx",
    )

    st.divider()
    st.subheader("Quick access")
    quick_cols = st.columns(4)
    quick_map = [("pareto_week", "📊 Pareto - Top 10"), ("action_plan", "🗂️ Action Plan"),
                 ("detailed", "📈 Detailed Graphs"), ("fault_code", "🎯 Fault Code Analysis")]
    for c, (key, label) in zip(quick_cols, quick_map):
        if c.button(label, use_container_width=True, key=f"quick_{key}"):
            st.session_state.page = key
            st.rerun()

# ---------------------------------------------------------------------------
# PAGE: PARETO - FAULT CODE
# ---------------------------------------------------------------------------
elif page == "fault_code":
    phase_banner(phase_label)
    st.header("🎯 Fault Code Analysis")
    st.caption("Ranking of fault codes by total duration, with cumulative % curve — dark theme.")

    pareto_fc = compute_pareto(df, group_col="Cause")
    fig = dark_pareto_fig(pareto_fc, "Cause", "Duree_totale_min", "Cumul_%", "Fault Code Analysis")
    st.plotly_chart(fig, use_container_width=True)

    with st.expander("View the Pareto level 1 table"):
        st.dataframe(pareto_fc, use_container_width=True)

    st.divider()
    st.subheader("Breakdown (%) by fault code")
    st.caption("Percentage shown outside the circle to stay readable even for small slices.")
    top_n_pie = st.slider("Number of fault codes shown (the rest are grouped into 'Other codes')",
                           min_value=3, max_value=15, value=8, key="pie_fc_topn")
    pie_data = pareto_fc[["Cause", "Duree_totale_min"]].copy()
    if len(pie_data) > top_n_pie:
        top_part = pie_data.iloc[:top_n_pie]
        rest = pd.DataFrame({"Cause": ["Other codes"], "Duree_totale_min": [pie_data.iloc[top_n_pie:]["Duree_totale_min"].sum()]})
        pie_data = pd.concat([top_part, rest], ignore_index=True)
    fig_donut = px.pie(pie_data, names="Cause", values="Duree_totale_min", hole=0.45,
                        color_discrete_sequence=VERSIGENT_PIE_COLORS)
    fig_donut.update_traces(textposition="outside", textinfo="percent+label", automargin=True)
    fig_donut.update_layout(
        paper_bgcolor=DARK_BG, plot_bgcolor=DARK_BG, font=dict(color="white"),
        height=520, showlegend=True, legend=dict(font=dict(color="white")),
        margin=dict(t=30, b=30, l=80, r=80),
    )
    st.plotly_chart(fig_donut, use_container_width=True)

    df_download_button("⬇️ Download the table (Excel)", pareto_fc, "pareto_fault_code.xlsx", "dl_fc")
    png = fig_to_png_bytes(fig)
    if png:
        st.download_button("🖼️ Download the chart (PNG)", png, file_name="pareto_fault_code.png",
                            mime="image/png", key="dl_fc_png")

# ---------------------------------------------------------------------------
# PAGE: PARETO OF THE WEEK (top 10)
# ---------------------------------------------------------------------------
elif page == "pareto_week":
    phase_banner(phase_label)
    st.header("📊 Pareto - Top 10")
    st.caption("The 10 most impactful pieces of equipment in the selected scope, by total failure duration.")

    pareto_week = compute_group_totals(df, group_col="Machine", top_n=10)
    fig = bar_fig(pareto_week, "Machine", "Valeur", "Pareto - Top 10", height=520)
    st.plotly_chart(fig, use_container_width=True)

    st.subheader("Top 10 list")
    listing = pareto_week.copy()
    listing.insert(0, "Rank", range(1, len(listing) + 1))
    listing = listing.rename(columns={"Machine": "Equipment", "Valeur": "Total duration (min)"})
    st.dataframe(listing, use_container_width=True, hide_index=True)

    df_download_button("⬇️ Download the table (Excel)", listing, "pareto_top10.xlsx", "dl_pareto_week")
    png = fig_to_png_bytes(fig)
    if png:
        st.download_button("🖼️ Download the chart (PNG)", png, file_name="pareto_top10.png",
                            mime="image/png", key="dl_pareto_week_png")

# ---------------------------------------------------------------------------
# PAGE: ISHIKAWA - TOP 10
# ---------------------------------------------------------------------------
elif page == "ishikawa":
    phase_banner(phase_label)
    st.header("🐟 Ishikawa Diagram - Top 10")
    st.caption(
        "Fishbone (Ishikawa) diagram built from the Pareto - Top 10: each of the 10 most "
        "impactful pieces of equipment becomes a branch pointing to the effect (total downtime)."
    )

    pareto_ishikawa = compute_group_totals(df, group_col="Machine", top_n=10)
    total_top10 = pareto_ishikawa["Valeur"].sum() if not pareto_ishikawa.empty else 0
    fig_ishi = build_ishikawa_fig(
        pareto_ishikawa, group_col="Machine", value_col="Valeur",
        effect_label=f"Total Downtime<br>Top 10<br>{total_top10:,.0f} min",
        title="Ishikawa Diagram - Top 10",
    )

    col_diag, col_table = st.columns([2, 1])
    with col_diag:
        st.plotly_chart(fig_ishi, use_container_width=True)
    with col_table:
        st.subheader("Top 10 list")
        listing_ishi = pareto_ishikawa.copy()
        listing_ishi.insert(0, "Rank", range(1, len(listing_ishi) + 1))
        listing_ishi = listing_ishi.rename(columns={"Machine": "Equipment", "Valeur": "Total duration (min)"})
        st.dataframe(listing_ishi, use_container_width=True, hide_index=True)

    df_download_button("⬇️ Download the table (Excel)", listing_ishi, "ishikawa_top10.xlsx", "dl_ishikawa")
    png_ishi = fig_to_png_bytes(fig_ishi)
    if png_ishi:
        st.download_button("🖼️ Download the diagram (PNG)", png_ishi, file_name="ishikawa_top10.png",
                            mime="image/png", key="dl_ishikawa_png")

# ---------------------------------------------------------------------------
# PAGE: PARETO BY EQUIPMENT (Level 1 & 2)
# ---------------------------------------------------------------------------
elif page == "pareto_equip":
    phase_banner(phase_label)
    st.header("🔍 Pareto by Equipment — Level 1 & 2")

    st.subheader("Pareto Level 1 — Failure causes (selected scope)")
    pareto1 = compute_pareto(df, group_col="Cause")
    fig1 = go.Figure()
    fig1.add_bar(x=pareto1["Cause"], y=pareto1["Duree_totale_min"], marker_color=VERSIGENT_ORANGE_HEX,
                 text=pareto1["Duree_totale_min"], textposition="outside")
    fig1.add_trace(go.Scatter(x=pareto1["Cause"], y=pareto1["Cumul_%"], name="Cumulative %",
                               yaxis="y2", mode="lines+markers", line=dict(color=VERSIGENT_BLACK_HEX)))
    fig1.add_hline(y=80, line_dash="dash", line_color="gray", yref="y2")
    fig1.update_layout(yaxis2=dict(title="Cumulative %", overlaying="y", side="right", range=[0, 100]),
                        xaxis_tickangle=-45, height=450)
    st.plotly_chart(fig1, use_container_width=True)
    with st.expander("View the table"):
        st.dataframe(pareto1, use_container_width=True)

    st.divider()
    st.subheader("Pareto Level 2 — Detail by machine for one cause")
    if not pareto1.empty:
        top_cause = st.selectbox("Choose the cause to break down", pareto1["Cause"].tolist())
        pareto2 = compute_pareto_level2(df, top_value=top_cause, level1_col="Cause", level2_col="Machine")
        if not pareto2.empty:
            fig2 = go.Figure()
            fig2.add_bar(x=pareto2["Machine"], y=pareto2["Duree_totale_min"], marker_color=VERSIGENT_ORANGE_HEX,
                         text=pareto2["Duree_totale_min"], textposition="outside")
            fig2.add_trace(go.Scatter(x=pareto2["Machine"], y=pareto2["Cumul_%"], name="Cumulative %",
                                       yaxis="y2", mode="lines+markers", line=dict(color=VERSIGENT_BLACK_HEX)))
            fig2.add_hline(y=80, line_dash="dash", line_color="gray", yref="y2")
            fig2.update_layout(yaxis2=dict(title="Cumulative %", overlaying="y", side="right", range=[0, 100]),
                                xaxis_tickangle=-45, height=450)
            st.plotly_chart(fig2, use_container_width=True)
            st.dataframe(pareto2, use_container_width=True)
        else:
            st.info("Not enough data for this cause.")

    st.divider()
    st.subheader("A Pareto + a pie chart per equipment (CM / ROB / Ultrasonic / Others)")
    tabs = st.tabs([e for e in EQUIPMENT_ORDER if e in df["Equipement"].unique()])
    for tab, equip in zip(tabs, [e for e in EQUIPMENT_ORDER if e in df["Equipement"].unique()]):
        with tab:
            full_data = equipement_data(df, equip)
            data = full_data.head(MAX_HOME_CHART_ITEMS)
            if data.empty:
                st.info("No data for this equipment in this scope.")
                continue
            if len(full_data) > MAX_HOME_CHART_ITEMS:
                st.caption(f"Showing top {MAX_HOME_CHART_ITEMS} of {len(full_data)} machines (see Data tab for full detail).")
            col_a, col_b = st.columns(2)
            col_a.plotly_chart(bar_fig(data, "Machine", "Valeur", f"Pareto — {equip}", height=420),
                                use_container_width=True)
            fig_pie = px.pie(data, names="Machine", values="Valeur", hole=0.4,
                              title=f"Breakdown (%) — {equip}", color_discrete_sequence=VERSIGENT_PIE_COLORS)
            fig_pie.update_traces(textposition="outside", textinfo="percent+label", automargin=True)
            fig_pie.update_layout(margin=dict(t=45, b=30, l=70, r=70))
            col_b.plotly_chart(fig_pie, use_container_width=True)
            st.dataframe(full_data, use_container_width=True)

    st.divider()
    st.subheader("Pareto by Failure Frequency")
    st.caption("Ranking by NUMBER of failures (not duration) — least reliable equipment.")
    freq_pareto = compute_group_totals(df, group_col="Machine", agg="count", top_n=MAX_HOME_CHART_ITEMS)
    fig_freq = bar_fig(freq_pareto, "Machine", "Valeur", "Failure frequency by equipment", color="#8C8C8C", height=450)
    st.plotly_chart(fig_freq, use_container_width=True)
    st.dataframe(freq_pareto, use_container_width=True)

# ---------------------------------------------------------------------------
# PAGE: DETAILED GRAPHS (dark theme, 4 tabs CM / ROB / Ultrasonic / Others)
# ---------------------------------------------------------------------------
elif page == "detailed":
    phase_banner(phase_label)
    st.header("📈 Detailed Graphs")
    st.caption("Equipment → failure title detail, for CM (+ Conveyor), ROB, Ultrasonic and Others.")

    task_col = "Task" if "Task" in df.columns else "Machine"
    detail_data = compute_detailed_breakdown(df, family_col="Equipement", machine_col="Machine", task_col=task_col)
    detail_data = {k: detail_data[k] for k in EQUIPMENT_ORDER if k in detail_data}

    if not detail_data:
        st.info("Not enough data for this scope.")
    else:
        tabs = st.tabs(list(detail_data.keys()))
        for tab, (equip, full_data) in zip(tabs, detail_data.items()):
            with tab:
                data = full_data.head(MAX_DETAIL_CHART_ITEMS)
                if len(full_data) > MAX_DETAIL_CHART_ITEMS:
                    st.caption(f"Showing top {MAX_DETAIL_CHART_ITEMS} of {len(full_data)} items by duration "
                               f"(full detail in the table below).")
                fig = go.Figure()
                fig.add_bar(
                    x=data["Duree_min"], y=[data["Machine"], data["Task"]],
                    orientation="h", marker_color=DARK_BAR,
                    text=data["Duree_min"], textposition="outside",
                    textfont=dict(color="white"),
                )
                fig.update_layout(
                    title=dict(text=equip, font=dict(color=DARK_ACCENT, size=20)),
                    paper_bgcolor=DARK_BG, plot_bgcolor=DARK_BG,
                    font=dict(color=DARK_ACCENT),
                    xaxis=dict(color="white", gridcolor=DARK_GRID),
                    yaxis=dict(color="white"),
                    height=max(420, 26 * len(data)),
                    margin=dict(l=10, r=10, t=50, b=10),
                )
                st.plotly_chart(fig, use_container_width=True)
                st.dataframe(full_data, use_container_width=True)

        st.divider()
        st.subheader("📦 Dedicated export — Detailed Graphs")
        st.caption("A standalone Excel file, just for this page (not mixed with the other exports).")
        buf_detail = BytesIO()
        with pd.ExcelWriter(buf_detail, engine="openpyxl") as writer:
            _write_detailed_graphs_workbook(writer, detail_data)
        st.download_button(
            "📦 Download Detailed Graphs (Excel, ready to send)",
            buf_detail.getvalue(),
            file_name=f"Detailed_Graphs_{phase_label_file}_{pd.Timestamp.today().date()}.xlsx",
            key="dl_detailed_xlsx",
        )

# ---------------------------------------------------------------------------
# PAGE: ACTION PLAN (exact Versigent format)
# ---------------------------------------------------------------------------
elif page == "action_plan":
    phase_banner(phase_label)
    st.header(f"🗂️ {phase_label} Action Plan")
    st.caption(
        "Automatically generated in the same format as the Versigent Excel action plan. "
        "Problem Description = « Completion Details » of the most recent failure ; "
        "Occurrence Date in WKxx-d format ; Cause is left BLANK for the operator to fill in ; "
        "Action Level = 3 ; Resp. = responsible engineer (reformatted) ; "
        "Effectivity Validation Method defaults to « Visual Control » ; "
        "validation = Ahmed BOUBEGHLI / Done."
    )
    if not completion_col:
        st.warning("⚠️ Enable the « Completion Details » column (section 2) to automatically fill Problem Description.")
    if not engineer_col:
        st.warning("⚠️ Enable the « Responsible Engineer » column (section 2) to automatically fill Resp.")

    top_n = st.slider("Number of pieces of equipment to include", min_value=3, max_value=20, value=8)
    plan = build_versigent_action_plan(df, machine_col="Machine", cause_col="Cause",
                                        completion_col=completion_col, date_col="Date",
                                        engineer_col=engineer_col, top_n=top_n)
    st.dataframe(plan, use_container_width=True)

    buf_plan = BytesIO()
    with pd.ExcelWriter(buf_plan, engine="openpyxl") as writer:
        _write_action_plan_sheet(writer, plan, phase_label)
    st.download_button("📦 Download the action plan (Excel, Versigent styling)", buf_plan.getvalue(),
                        file_name=f"Action_Plan_Versigent_{pd.Timestamp.today().date()}.xlsx",
                        key="dl_action_plan")

# ---------------------------------------------------------------------------
# PAGE: VIEW BY DATE
# ---------------------------------------------------------------------------
elif page == "by_date":
    phase_banner(phase_label)
    st.header("📅 View by Date — failures on a specific day")
    all_dates = sorted(df["Date"].dropna().dt.date.unique())
    if not all_dates:
        st.info("No valid dates in the selected scope.")
    else:
        view_mode = st.radio(
            "View mode", ["Single day", "Date range"], horizontal=True,
            help="Single day: pick one calendar date. Date range: pick a start and an "
                 "end date (e.g. 9/6 to 13/6/2026) and the dashboard combines every "
                 "failure in between automatically.",
        )

        if view_mode == "Single day":
            chosen_date = st.date_input(
                "Choose a date", value=all_dates[-1],
                min_value=all_dates[0], max_value=all_dates[-1],
            )
            day_df = df[df["Date"].dt.date == chosen_date]
            period_label = f"{chosen_date}"
            period_label_file = f"{chosen_date}"
        else:
            dcol1, dcol2 = st.columns(2)
            start_date = dcol1.date_input(
                "From", value=all_dates[0], min_value=all_dates[0], max_value=all_dates[-1], key="range_start_date",
            )
            end_date = dcol2.date_input(
                "To", value=all_dates[-1], min_value=all_dates[0], max_value=all_dates[-1], key="range_end_date",
            )
            if start_date > end_date:
                st.warning("⚠️ The start date is after the end date — please adjust the range.")
                day_df = df.iloc[0:0]
            else:
                day_df = df[(df["Date"].dt.date >= start_date) & (df["Date"].dt.date <= end_date)]
            period_label = f"{start_date} to {end_date}"
            period_label_file = f"{start_date}_to_{end_date}"

        if day_df.empty:
            st.success(f"✅ No failure recorded on {period_label}.")
        else:
            c1, c2, c3 = st.columns(3)
            c1.metric("Failures in period" if view_mode == "Date range" else "Failures that day", len(day_df))
            c2.metric("Total downtime", f"{round(day_df['Duree_min'].sum() / 60, 1)} h")
            c3.metric("Most affected machine", day_df.groupby("Machine")["Duree_min"].sum().idxmax())

            show_cols = [c for c in ["Date", "Machine", "Equipement", "Cause", "Task", "Duree_min"] if c in day_df.columns]
            st.dataframe(day_df[show_cols].rename(columns={"Equipement": "Equipment"}).sort_values("Duree_min", ascending=False),
                         use_container_width=True)

            pareto_day = compute_group_totals(day_df, group_col="Machine")
            fig = bar_fig(pareto_day, "Machine", "Valeur", f"Failures on {period_label}", height=420)
            st.plotly_chart(fig, use_container_width=True)

            df_download_button("⬇️ Download this period's failures (Excel)",
                                day_df[show_cols], f"failures_{period_label_file}.xlsx", "dl_day")

# ---------------------------------------------------------------------------
# PAGE: TREND & ALERTS
# ---------------------------------------------------------------------------
elif page == "trend":
    phase_banner(phase_label)
    st.header("📈 Paynter Chart — Trend of Causes")
    paynter = compute_paynter(df, group_col="Cause")
    if not paynter.empty and paynter.shape[1] > 0:
        fig3 = go.Figure()
        for cause in paynter.index:
            fig3.add_trace(go.Scatter(x=paynter.columns, y=paynter.loc[cause], mode="lines+markers", name=cause))
        fig3.update_layout(xaxis_title="Week", yaxis_title="Duration (min)", height=450)
        st.plotly_chart(fig3, use_container_width=True)
        with st.expander("View the Paynter table"):
            st.dataframe(paynter, use_container_width=True)
    else:
        st.info("Not enough different weeks in the data to plot a Paynter chart.")

    st.divider()
    st.header("🚨 Critical Equipment Alerts")
    threshold = st.slider("Alert threshold (cumulative minutes over the period)", min_value=10, max_value=500, value=100, step=10)
    alerts = compute_alerts(df, threshold_min=threshold, group_col="Machine")
    if not alerts.empty:
        st.error(f"{len(alerts)} piece(s) of equipment exceed the {threshold} min threshold:")
        st.dataframe(alerts, use_container_width=True)
    else:
        st.success("No equipment exceeds the current threshold.")

    st.divider()
    st.header("🔮 Prediction — Likely Upcoming Downtime")
    if not paynter.empty:
        cause_pred = st.selectbox("Choose the cause to predict", paynter.index.tolist(), key="pred")
        result = predict_next_week(paynter, cause_pred)
        if result["prediction"] is not None:
            st.metric(f"Prediction for \u00ab {cause_pred} \u00bb", f"{result['prediction']} min", help=f"Method: {result['methode']}")
        else:
            st.info("Not enough data to predict.")
    else:
        st.info("Not enough data for a prediction.")

# ---------------------------------------------------------------------------
# PAGE: DATA
# ---------------------------------------------------------------------------
elif page == "data":
    phase_banner(phase_label)
    st.header("🗄️ Data")
    st.caption(f"Current scope: Phase = {selected_phase} · Equipment = "
               f"{', '.join(selected_equipements) if selected_equipements else 'all'} · {len(df)} rows "
               f"(week {EXCLUDED_WEEK} excluded).")
    show_cols = [c for c in ["Date", "Machine", "Equipement", "Phase", "Cause", "Task", "Duree_min"] if c in df.columns]
    st.dataframe(df[show_cols].rename(columns={"Equipement": "Equipment"}).sort_values("Date"),
                 use_container_width=True, height=500)

    st.subheader("Breakdown by Equipment")
    st.dataframe(df.groupby("Equipement")["Duree_min"].agg(["count", "sum"]).rename(
        columns={"count": "Nb failures", "sum": "Total duration (min)"}), use_container_width=True)

    df_download_button("⬇️ Download the scope's data (Excel)", df[show_cols],
                        f"data_{pd.Timestamp.today().date()}.xlsx", "dl_data_full")

# ---------------------------------------------------------------------------
# PAGE: GLOBAL EXPORT (Home + Pareto - Top 10 + Action Plan + Data)
# ---------------------------------------------------------------------------
elif page == "export":
    phase_banner(phase_label)
    st.header("📦 Global Export")
    st.caption(
        "A single Excel file ready to send, with one page per section (Assembly/Cutting Analysis, "
        "Pareto - Top 10, Ishikawa - Top 10, Action Plan, Data, Detailed Graphs) — well organized, "
        "large legible text, Versigent colors and logo. (A standalone, per-equipment version of "
        "Detailed Graphs is also available on its own page.)"
    )

    kpis = summary_kpis(df)
    pareto1 = compute_pareto(df, group_col="Cause")
    resume_text = generate_text_summary(df, pareto1, group_col="Cause")
    data_cm = equipement_data(df, "CM", top_n=MAX_HOME_CHART_ITEMS)
    data_rob = equipement_data(df, "ROB", top_n=MAX_HOME_CHART_ITEMS)
    data_us = equipement_data(df, "Ultrasonic", top_n=MAX_HOME_CHART_ITEMS)
    data_others = equipement_data(df, "Others", top_n=MAX_HOME_CHART_ITEMS)
    pareto_week_top10 = compute_group_totals(df, group_col="Machine", top_n=10)
    plan = build_versigent_action_plan(df, machine_col="Machine", cause_col="Cause",
                                        completion_col=completion_col, date_col="Date",
                                        engineer_col=engineer_col, top_n=10)
    show_cols = [c for c in ["Date", "Machine", "Equipement", "Phase", "Cause", "Task", "Duree_min"] if c in df.columns]

    task_col_export = "Task" if "Task" in df.columns else "Machine"
    detail_data_export = compute_detailed_breakdown(df, family_col="Equipement", machine_col="Machine", task_col=task_col_export)
    detail_data_export = {k: detail_data_export[k] for k in EQUIPMENT_ORDER if k in detail_data_export}

    export_home_title = "Global Analysis" if (phase_label.startswith("All") or phase_label == "Global") else f"{phase_label} Analysis"

    buf_global = BytesIO()
    with pd.ExcelWriter(buf_global, engine="openpyxl") as writer:
        _write_home_sheet(writer, df, phase_label, week_label, kpis, resume_text,
                           data_cm, data_rob, data_us, data_others, sheet_name=f"1. {export_home_title}")
        _write_pareto_week_sheet(writer, pareto_week_top10, week_label)
        _write_ishikawa_sheet(writer, pareto_week_top10)
        _write_action_plan_sheet(writer, plan, phase_label, sheet_name="4. Action Plan")
        _write_data_sheet(writer, df[show_cols])
        _write_detailed_graphs_summary_sheet(writer, detail_data_export)

    st.download_button(
        "📦 Download the global Excel report (6 pages)",
        buf_global.getvalue(),
        file_name=f"Versigent_Report_{phase_label_file}_{pd.Timestamp.today().date()}.xlsx",
        key="dl_export_global_xlsx",
    )

    st.divider()
    st.subheader("Complete PowerPoint report")
    with st.expander("✏️ Customize the presentation text"):
        pptx_title = st.text_input("Title (cover slide)", value="Versigent")
        pptx_subtitle = st.text_input("Subtitle", value=f"{phase_label}")
        pptx_closing = st.text_area("Closing slide text (leave empty to skip it)", value="")

    fig_week = bar_fig(pareto_week_top10, "Machine", "Valeur", "Pareto - Top 10")
    sections = [
        {
            "title": "Summary",
            "kpis": [
                ("Number of failures", kpis["nb_evenements"]),
                ("Total downtime", f"{kpis['duree_totale_h']} h"),
                ("Average", f"{kpis['duree_moyenne_min']} min"),
                ("Most critical equipment", kpis["machine_top"]),
                ("Most frequent cause", kpis["cause_top"]),
            ],
            "body_text": resume_text,
        },
        {"title": "📊 Pareto - Top 10", "image": fig_to_png_bytes(fig_week), "table": pareto_week_top10},
    ]
    for equip, data in [("CM", data_cm), ("ROB", data_rob), ("Ultrasonic", data_us), ("Others", data_others)]:
        if data.empty:
            continue
        fig_e = bar_fig(data, "Machine", "Valeur", f"Pareto — {equip}")
        sections.append({"title": f"🏭 Analysis — {equip}", "image": fig_to_png_bytes(fig_e), "table": data})
    sections.append({
        "title": "🗂️ Action Plan",
        "caption": "Automatically generated from the most impactful equipment in the scope.",
        "table": plan,
        "table_header_color": ACCENT_COLOR,
    })

    buf_pptx = build_pptx_report(
        sections, title=pptx_title, subtitle=pptx_subtitle if pptx_subtitle else None,
        closing_text=pptx_closing if pptx_closing.strip() else None, logo_bytes=LOGO_BYTES,
    )
    st.download_button(
        "📦 Download the complete PowerPoint report",
        buf_pptx.getvalue(),
        file_name=f"Versigent_Downtime_Report_{phase_label_file}_{pd.Timestamp.today().date()}.pptx",
        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        key="dl_export_pptx",
    )

st.divider()
st.caption(
    "Developed as part of a Final Year Project — Mechatronics Engineering. "
    "Based on real Versigent downtime tracking data (Kaizen Event 4Q)."
)
